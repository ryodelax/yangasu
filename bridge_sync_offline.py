import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from bridge_live_refresh import (
    BANK_ROOT,
    MATCH_STATUS_UNMATCHED,
    apply_summary_corrections,
    bank_dedupe_key_from_record,
    bank_display_summary,
    build_bank_records_from_csv,
    build_direction_rows,
    build_payment_progress,
    build_receivable_forecast,
    build_sales_rows,
    chunked_rows,
    date_key,
    ensure_row_width,
    get_header_map,
    get_range_map,
    history_rows_existing,
    import_bank_files,
    load_bank_import_rules,
    merge_bank_rows,
    run_matching,
    safe_string,
)


ROOT = Path("/Users/hasegawaryou11/Downloads/Cursol")
RAW_DIR = ROOT / "_live_raw"
OUT_DIR = ROOT / "_generated_sync"


def sheet_values(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else value for value in row]
        while values and values[-1] == "":
            values.pop()
        rows.append(values)
    while rows and not any(str(v).strip() for v in rows[-1]):
        rows.pop()
    return rows


def workbook_sheet_map(path):
    wb = load_workbook(path, data_only=False)
    return {name: sheet_values(wb[name]) for name in wb.sheetnames}


def write_tsv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        for row in rows:
            fh.write("\t".join("" if value is None else str(value) for value in row))
            fh.write("\n")


def move_file_safe(src, dst_dir):
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / src.name
    if dst.exists():
        dst = dst_dir / f"{src.stem}__{datetime.now().strftime('%Y%m%d%H%M%S')}{src.suffix}"
    shutil.move(str(src), str(dst))
    return str(dst)


def import_bank_files_offline(bank_rows, correction_rows, history_rows, dry_run=True):
    existing_files = history_rows_existing(history_rows)
    bank_header = bank_rows[0]
    header_width = len(bank_header)
    existing_records = []
    for raw in bank_rows[1:]:
        row = ensure_row_width(raw, header_width)
        existing_records.append({bank_header[i]: row[i] for i in range(header_width)})
    existing_keys = {bank_dedupe_key_from_record(rec): True for rec in existing_records}
    rules = load_bank_import_rules(correction_rows)
    imported = []
    history_add = []
    result_moves = {"success": [], "error": []}
    scanned = []
    pending_moves = {"success": [], "error": []}

    for path in sorted(BANK_ROOT.iterdir()):
        if path.is_dir() or path.suffix.lower() != ".csv":
            continue
        scanned.append(path.name)
        is_already_logged = path.name in existing_files
        records, error = build_bank_records_from_csv(path, rules)
        imported_at = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        process_id = datetime.now().strftime("%Y%m%d%H%M%S") + "_" + path.name.replace(" ", "_")
        if error:
            if not is_already_logged:
                history_add.append([process_id, imported_at, path.name, "", 0, 0, 0, 1, "失敗", error, "エラー", "再実行待ち"])
            pending_moves["error"].append((str(path), str(BANK_ROOT / "エラー")))
            if not dry_run:
                result_moves["error"].append(move_file_safe(path, BANK_ROOT / "エラー"))
            continue

        new_count = 0
        skip_count = 0
        for record in records:
            key = bank_dedupe_key_from_record(record)
            if key and key in existing_keys:
                skip_count += 1
                continue
            existing_keys[key] = True
            record["取込処理ID"] = process_id
            record["取込ファイル名"] = path.name
            record["取込日時"] = imported_at
            imported.append(record)
            new_count += 1

        if not is_already_logged:
            history_add.append([process_id, imported_at, path.name, "", len(records), new_count, skip_count, 0, "成功", "", "読み取り済み", ""])
        pending_moves["success"].append((str(path), str(BANK_ROOT / "読み取り済み")))
        if not dry_run:
            result_moves["success"].append(move_file_safe(path, BANK_ROOT / "読み取り済み"))

    merged_rows = merge_bank_rows(existing_records, imported, bank_header)
    return merged_rows, history_add, result_moves, scanned, pending_moves


def build_sync_payload(dry_run=True):
    main = workbook_sheet_map(RAW_DIR / "main.xlsx")
    source = workbook_sheet_map(RAW_DIR / "source.xlsx")

    sales_rows = build_sales_rows(
        {
            "顧客対応状況（車検）": source.get("顧客対応状況（車検）", []),
            "顧客対応状況（12点）": source.get("顧客対応状況（12点）", []),
            "顧客対応状況（車両販売）": source.get("顧客対応状況（車両販売）", []),
        },
        main["振込入金リスト一覧"],
    )

    recurring_rows = main.get("定期入金リスト")
    if not recurring_rows:
        recurring_rows = [["有効", "摘要", "金額", "業務No.", "顧客No.", "ステータス", "確認メモ"]]

    merged_bank_rows, history_add, moved, scanned, pending_moves = import_bank_files_offline(
        main["銀行データチェック用"],
        main["銀行文字補正学習"],
        main["取込履歴"],
        dry_run=dry_run,
    )

    merged_bank_rows, sales_rows, review_memos = run_matching(
        merged_bank_rows,
        sales_rows,
        main["※名義対応表"],
        main["照合学習マスタ"],
        recurring_rows,
    )
    sales_rows = build_payment_progress(merged_bank_rows, sales_rows)
    forecast_rows = build_receivable_forecast(sales_rows)
    deposit_rows, withdrawal_rows = build_direction_rows(merged_bank_rows)

    existing_history = main["取込履歴"]
    history_header = existing_history[0] if existing_history else ["処理ID", "取込日時", "ファイル名", "fileId", "対象行数", "成功行数", "スキップ行数", "失敗行数", "結果", "失敗理由", "移動先", "再実行状況"]
    history_body = existing_history[1:] if len(existing_history) > 1 else []
    history_rows = [history_header] + history_body + history_add

    review_header = main["銀行要確認メモ"][0] if main.get("銀行要確認メモ") else ["日付", "摘要", "金額", "業務No候補", "顧客No候補", "確認メモ"]
    review_rows = [review_header] + review_memos

    outputs = {
        "銀行データチェック用": merged_bank_rows,
        "振込入金リスト一覧": sales_rows,
        "売掛入金見込み管理": forecast_rows,
        "取込履歴": history_rows,
        "銀行要確認メモ": review_rows,
        "銀行入金一覧": deposit_rows,
        "銀行出金一覧": withdrawal_rows,
    }
    for sheet, rows in outputs.items():
        write_tsv(OUT_DIR / f"{sheet}.tsv", rows)

    summary = {
        "scanned": scanned,
        "pending_moves": pending_moves,
        "bank_rows": len(merged_bank_rows) - 1,
        "sales_rows": len(sales_rows) - 1,
        "forecast_rows": len(forecast_rows) - 1,
        "review_rows": len(review_rows) - 1,
        "deposit_rows": len(deposit_rows) - 1,
        "withdrawal_rows": len(withdrawal_rows) - 1,
    }
    (OUT_DIR / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


if __name__ == "__main__":
    summary = build_sync_payload(dry_run=True)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
